import openpyxl
import json
from pprint import pprint
jsondata = open("路外停車資訊.json", 'r', encoding='UTF-8')
data = jsondata.read()
data = json.loads(data)
seta = ('areaId', 'areaName', 'parkName', 'totalSpace', 'surplusSpace', 'payGuide', 'introduction', 'address', 'wgsX', 'wgsY', 'parkId')
seta2 = ('A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1')

workbook = openpyxl.Workbook()
workbook.remove_sheet(workbook.get_sheet_by_name('Sheet'))
workbook.create_sheet('info')
sheet = workbook.get_sheet_by_name('info')

x = 0
while x < len(seta):
    sheet[seta2[x]] = seta[x]
    x+=1



x = 0
w = 1
while x < len(seta):

    y = 0
    z = 2
    while y < len(data['parkingLots']):
        
        sheet.cell(row = z, column = w).value = (data['parkingLots'][y][seta[x]])
        y+=1
        z+=1
        
    x+=1
    w+=1
   



workbook.save('parkinginfotest.xlsx')

